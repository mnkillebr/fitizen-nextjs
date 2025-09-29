import time
from typing import Annotated
from fastapi import APIRouter, Form, Query
from fastapi.responses import StreamingResponse, FileResponse
import asyncio
import json
import pandas as pd
import io
import tempfile
import os

from agents.models.profile import Client, FitnessProfile
from agents.crews.parq_program_crew.parq_program_crew import client, fitness_profile, movement_patterns, movement_plane, balance_type, parq_program_crew, my_listener
from agents.flows.generate_program_flow.generate_program_flow import GenerateProgramFlow, test_fms, program_flow_listener
from agents.models.program import GenerateProgramInput

router = APIRouter(
    prefix="/programs",
    tags=["programs"]
)

def convert_program_to_excel(program_data: dict):
    """
    Convert program data to Excel format matching the workout template structure
    """
    # Create a BytesIO object to store the Excel file
    excel_buffer = io.BytesIO()

    program_summary = program_data['program_summary']
    full_program = program_data['full_program']
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # Create summary sheet
        summary_rows = []
        summary_rows.append(['FITNESS PROGRAM SUMMARY'])
        summary_rows.append([program_summary])
        summary_rows.append([])
        summary_rows.append(['Week', 'Days', 'Focus Areas'])
        
        for week_key, week_data in full_program.items():
            week_name = week_key.replace('week', 'Week ')
            days = len(week_data)
            focus_areas = []
            
            # Extract focus areas from the week
            for day_data in week_data:
                # Convert Pydantic model to dict if needed
                if hasattr(day_data, 'model_dump'):
                    day_data = day_data.model_dump()
                
                if day_data.get('power'):
                    focus_areas.append('Power')
                if day_data.get('circuit_1') or day_data.get('circuit_2'):
                    focus_areas.append('Strength/Circuit')
                if day_data.get('finisher'):
                    focus_areas.append('Conditioning')
            
            focus_areas = list(set(focus_areas))  # Remove duplicates
            summary_rows.append([week_name, days, ', '.join(focus_areas)])
        
        # Write summary sheet
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
        
        # Format summary sheet
        summary_worksheet = writer.sheets['Summary']
        for column in summary_worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Process each week
        for week_key, week_data in full_program.items():
            week_name = week_key.replace('week', 'Week ')
            
            # Process each day in the week
            for day_idx, day_data in enumerate(week_data):
                # Convert Pydantic model to dict if needed
                if hasattr(day_data, 'model_dump'):
                    day_data = day_data.model_dump()
                
                day_name = day_data.get('day', 'Unknown Day')
                sheet_name = f"{week_name} - Day {day_name}"
                
                # Create the workout template structure
                workout_rows = []
                
                # Header section
                workout_rows.append(['Date:', '', '', 'Routine:', f'Day {day_name}', ''])
                # workout_rows.append([])  # Empty row
                
                # Correctives/Movement Prep section
                workout_rows.append(['Correctives/Movement Prep', '', '','Notes', '', ''])
                movement_prep = day_data.get('movement_prep', [])
                if movement_prep:
                    for prep in movement_prep:
                        # Convert Pydantic model to dict if needed
                        if hasattr(prep, 'model_dump'):
                            prep = prep.model_dump()
                        
                        # Add foam rolling
                        foam_rolling = prep.get('foam_rolling', [])
                        if foam_rolling:
                            workout_rows.append(['foam roll', ', '.join(foam_rolling), '', '', '', ''])
                        
                        # Add dynamic stretches
                        dynamic_stretches = prep.get('dynamic_stretches', [])
                        if dynamic_stretches:
                            workout_rows.append(['dynamic stretches', ', '.join(dynamic_stretches), '', '', '', ''])
                        
                        # Add activation exercises
                        activation_exercises = prep.get('activation_exercises', [])
                        if activation_exercises:
                            workout_rows.append(['activation', ', '.join(activation_exercises), '', '', '', ''])
                else:
                    # Add placeholder rows
                    for _ in range(3):
                        workout_rows.append(['', '', '', '', '', ''])
                
                # workout_rows.append([])  # Empty row
                
                # Power/Speed/Agility Training section
                workout_rows.append(['Power/Speed/Agility Training', '', 'Notes', 'Sets/Reps/Time', 'Load', 'Rest'])
                power = day_data.get('power', '')
                if power:
                    workout_rows.append([power, '', '', '', '', ''])
                else:
                    workout_rows.append(['', '', '', '', '', ''])
                
                # Add more empty rows for power section
                for _ in range(3):
                    workout_rows.append(['', '', '', '', '', ''])
                
                # workout_rows.append([])  # Empty row
                
                # Resistance Training section
                workout_rows.append(['Resistance Training', '', 'Notes', 'Sets/Reps/Time', 'Load', 'Rest'])
                
                # Circuit 1
                circuit_1 = day_data.get('circuit_1', {})
                if circuit_1:
                    # Convert Pydantic model to dict if needed
                    if hasattr(circuit_1, 'model_dump'):
                        circuit_1 = circuit_1.model_dump()
                    
                    exercises = circuit_1.get('exercises', [])
                    for exercise in exercises:
                        # Convert Pydantic model to dict if needed
                        if hasattr(exercise, 'model_dump'):
                            exercise = exercise.model_dump()
                        
                        workout_rows.append([
                            exercise.get('name', ''),
                            '',
                            f"Circuit 1 - RPE {exercise.get('rpe', '')}",
                            f"{circuit_1.get('rounds', '')} rounds x {exercise.get('reps', '')} reps",
                            '',
                            f"{circuit_1.get('rest', '')}s rest"
                        ])
                
                # Circuit 2
                circuit_2 = day_data.get('circuit_2', {})
                if circuit_2:
                    # Convert Pydantic model to dict if needed
                    if hasattr(circuit_2, 'model_dump'):
                        circuit_2 = circuit_2.model_dump()
                    
                    exercises = circuit_2.get('exercises', [])
                    for exercise in exercises:
                        # Convert Pydantic model to dict if needed
                        if hasattr(exercise, 'model_dump'):
                            exercise = exercise.model_dump()
                        
                        workout_rows.append([
                            exercise.get('name', ''),
                            '',
                            f"Circuit 2 - RPE {exercise.get('rpe', '')}",
                            f"{circuit_2.get('rounds', '')} rounds x {exercise.get('reps', '')} reps",
                            '',
                            f"{circuit_2.get('rest', '')}s rest"
                        ])
                
                # Add more empty rows for resistance training
                resistance_exercises = [row for row in workout_rows if len(row) > 0 and row[0] and 'Resistance Training' not in row[0] and 'Notes' not in row[0] and 'Correctives' not in row[0] and 'Power' not in row[0] and 'Energy' not in row[0]]
                remaining_rows = 10 - len(resistance_exercises)
                for _ in range(max(0, remaining_rows)):
                    workout_rows.append(['', '', '', '', '', ''])
                
                # workout_rows.append([])  # Empty row
                
                # Energy System Development section
                workout_rows.append(['Energy System Development', '','Notes', 'Volume/Intensity/Rest', '', ''])
                finisher = day_data.get('finisher', '')
                if finisher:
                    workout_rows.append([finisher, '', '', '', '', ''])
                else:
                    workout_rows.append(['', '', '', '', '', ''])
                
                # Add more empty rows for energy system development
                for _ in range(2):
                    workout_rows.append(['', '', '', '', '', ''])
                
                # Create DataFrame and write to Excel
                df = pd.DataFrame(workout_rows)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting to match the template
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                
                # Define styles
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                thick_border = Border(
                    left=Side(style='thick'),
                    right=Side(style='thick'),
                    top=Side(style='thick'),
                    bottom=Side(style='thick')
                )
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply formatting to cells
                for row_idx, row in enumerate(workout_rows, 1):
                    for col_idx, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                        cell.border = border
                        
                        # Apply header formatting to section headers
                        if cell_value in ['Correctives/Movement Prep', 'Power/Speed/Agility Training', 'Resistance Training', 'Energy System Development']:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                        
                        # Apply header formatting to column headers
                        elif cell_value in ['Notes', 'Sets/Reps/Time', 'Load', 'Rest', 'Volume/Intensity/Rest']:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                
                # Apply cell merging after all values are set
                for row_idx, row in enumerate(workout_rows, 1):
                    for col_idx, cell_value in enumerate(row, 1):
                        # Merge cells for section headers
                        if cell_value in ['Correctives/Movement Prep', 'Power/Speed/Agility Training', 'Resistance Training', 'Energy System Development', 'Volume/Intensity/Rest']:
                            worksheet.merge_cells(f'{get_column_letter(col_idx)}{row_idx}:{get_column_letter(col_idx+1)}{row_idx}')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Reset buffer position
    excel_buffer.seek(0)
    return excel_buffer

async def parq_program_event_generator():
    """Generate SSE events from the listener queue"""
    while True:
        event = my_listener.get_event()
        if event:
            if event["type"] == "crew_completed":
                yield f"data: {json.dumps(event)}\n\n"
                break
            yield f"data: {json.dumps(event)}\n\n"
        await asyncio.sleep(0.1)

@router.get("/parq_program/events")
async def get_crew_events():
    """SSE endpoint for getting crew execution events"""
    return StreamingResponse(
        parq_program_event_generator(),
        media_type="text/event-stream"
    )

@router.post("/parq_program")
async def getParQProgram(profileClient: Annotated[Client, Query()], profileData: Annotated[FitnessProfile, Form()]):
    start_time = time.perf_counter()
    
    # print("profile client", profileClient.model_dump())
    # print("profile data", profileData.model_dump())
    # Clear any previous events
    my_listener.clear_events()
    
    # Prepare inputs
    crew_inputs = {
        'client': profileClient.model_dump(),
        'fitness_profile': profileData.model_dump(),
        # 'client': client,
        # 'fitness_profile': fitness_profile,
        'movement_patterns': movement_patterns,
        'movement_plane': movement_plane,
        'balance_type': balance_type,
    }
    # Start the crew execution
    result = await parq_program_crew.kickoff_async(
        inputs=crew_inputs
    )
    raw_output = result.raw
    # pydantic_output = result.pydantic.model_dump()
    process_time = time.perf_counter() - start_time
    
    return {
        "process_time": process_time,
        # "pydantic_output": pydantic_output,
        "raw_output": raw_output,
    }

async def program_flow_event_generator():
    """Generate SSE events from the listener queue"""
    while True:
        event = program_flow_listener.get_event()
        print("program flow event", event)
        if event:
            if event["type"] == "flow_finished":
                yield f"data: {json.dumps(event)}\n\n"
                break
            # Stream all intermediate events (including method_started/finished, crew/agent events)
            yield f"data: {json.dumps(event)}\n\n"
        await asyncio.sleep(0.1)

@router.get("/program_flow/events")
async def get_program_flow_events():
    """SSE endpoint for getting crew execution events"""
    return StreamingResponse(
        program_flow_event_generator(),
        media_type="text/event-stream",
    )

@router.post("/program_flow")
async def program_flow(programInput: Annotated[GenerateProgramInput, Form()], format: Annotated[str | None, Query()] = "json"):
    """Test the flow and return either JSON or Excel file"""
    start_time = time.perf_counter()
    # Test FMS Inputs
    # Clear any previous events
    program_flow_listener.clear_events()

    program_input = programInput.model_dump()
    # fms_input = test_fms
    fms_input = {
        'deep_squat': program_input['deepSquat'],
        'hurdle_step': program_input['hurdleStep'],
        'inline_lunge': program_input['inlineLunge'],
        'shoulder_mobility': program_input['shoulderMobility'],
        'active_straight_leg_raise': program_input['activeStraightLegRaise'],
        'trunk_stability_pushup': program_input['trunkStabilityPushUp'],
        'rotary_stability': program_input['rotaryStability'],
        'total_score': program_input['deepSquat'] + program_input['hurdleStep'] + program_input['inlineLunge'] + program_input['shoulderMobility'] + program_input['activeStraightLegRaise'] + program_input['trunkStabilityPushUp'] + program_input['rotaryStability']
    }
    coach_notes = program_input['coachNotes']

    print(f"FMS Input: {fms_input}")

    flow = GenerateProgramFlow(fms=fms_input, coach_notes=coach_notes)
    # flow.plot()
    result = await flow.kickoff_async()

    process_time = time.perf_counter() - start_time

    # Return JSON format
    if format.lower() == "json":
        return {
            "process_time": process_time,
            "result": result,
        }
    
    # Return Excel format
    elif format.lower() == "excel":
        # Convert result to Excel format
        excel_buffer = convert_program_to_excel(result)
        
        # Create a temporary file to store the Excel data
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(excel_buffer.getvalue())
            tmp_file_path = tmp_file.name

        # Return the Excel file
        return FileResponse(
            path=tmp_file_path,
            filename=f"fitness_program_default.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    else:
        return {"error": "Invalid format. Use 'json' or 'excel'"}

@router.post("/program_flow/excel")
async def program_flow_excel(programInput: Annotated[GenerateProgramInput, Form()]):
    """Test the flow and return Excel file directly"""
    start_time = time.perf_counter()
    # Clear any previous events
    program_flow_listener.clear_events()

    # Test FMS Inputs
    program_input = programInput.model_dump()
    # fms_input = test_fms
    fms_input = {
        'deep_squat': program_input['deepSquat'],
        'hurdle_step': program_input['hurdleStep'],
        'inline_lunge': program_input['inlineLunge'],
        'shoulder_mobility': program_input['shoulderMobility'],
        'active_straight_leg_raise': program_input['activeStraightLegRaise'],
        'trunk_stability_pushup': program_input['trunkStabilityPushUp'],
        'rotary_stability': program_input['rotaryStability'],
        'total_score': program_input['deepSquat'] + program_input['hurdleStep'] + program_input['inlineLunge'] + program_input['shoulderMobility'] + program_input['activeStraightLegRaise'] + program_input['trunkStabilityPushUp'] + program_input['rotaryStability']
    }
    coach_notes = program_input['coachNotes']

    print(f"FMS Input: {fms_input}")

    flow = GenerateProgramFlow(fms=fms_input, coach_notes=coach_notes)
    # flow.plot()
    result = await flow.kickoff_async()

    process_time = time.perf_counter() - start_time

    # Convert result to Excel format
    excel_buffer = convert_program_to_excel(result)
    
    # Create a temporary file to store the Excel data
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(excel_buffer.getvalue())
        tmp_file_path = tmp_file.name

    # Return the Excel file
    return FileResponse(
        path=tmp_file_path,
        filename=f"fitness_program_default.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )